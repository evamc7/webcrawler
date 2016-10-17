import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

#get data from a url response as json

# Find if the file already exists

if os.path.isfile("Users/IAMIN/Desktop/list_of_inventors_2016.xlsx") == True:
    wb = load_workbook("/Users/IAMIN/Desktop/list_of_inventors_2016.xlsx")
else:
    wb = Workbook("/Users/IAMIN/Desktop/list_of_inventors_2016.xlsx")

months = {
    "1": January, "2": February, "3": March, "4": April, "5": May, "6": June, "7": July, "8": August, "9": September, "10": October, "11": November, "12": December,
}

# if newmonth create new sheet, else write all the data in the same shee
d1 = datetime.today()

#find and format the month 

# If the info is from the same month add that to that sheet, else create a new one
if #date.today != data.month:
    new_sheet = wb.create_sheet(month)
    
    #Create the headings in the csv file

    new_sheet["A1"] = "Patent Number"
    new_sheet["B1"] = "Title"
    new_sheet["C1"] = "Publication date"
    new_sheet["D1"] = "For who"
    new_sheet["E1"] = "Inventor"

    #loop your data and put it in each cell
    for row in range(1, len(data)): #where data is the lenght of the json
        for col in range(0, 5):
            _ = new_sheet.cell(column=col, row=row, value="{0}".data)

    
else:
    sheet = wb[month] #where month is the sheet name which must be a month name

    #find last written row, print from that one +1 


wb.save("/Users/IAMIN/Desktop/list_of_inventors_2016.xlsx", as_template=False)